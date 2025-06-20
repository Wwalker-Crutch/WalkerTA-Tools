"""
Name:
    ExcelHandler
Author:
    William Walker @ Crutchfield
Description:

imports:
    @WalkerLog.py
    termcolor   ---> For colored console output
"""
from WalkerLog import *
from termcolor import colored
import GLOBALS
import os
from openpyxl import load_workbook
from datetime import datetime
import win32ui
from SpecsFileParser import EmailSpecsFromFile


def GetTemplatePath():
    base_dir = os.path.dirname(os.path.abspath(__file__))

    project_root = os.path.join(base_dir, "..")
    template_path = os.path.normpath(os.path.join(project_root, "Program", "Files", "WalkerTA-MISP-Template.xlsx"))

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found at: {template_path}")

    return template_path


def MakeWalkerExcelFolder():
    walker_root = MakeWalkerToolsFolder()

    excel_folder = os.path.join(walker_root, "WalkerTAExcel")
    os.makedirs(excel_folder, exist_ok=True)

    log(f"\n[EXCEL_FOLDER] Created (or confirmed) Excel folder at: {excel_folder}")
    return excel_folder


def SelectEMLFile():
    dlg = win32ui.CreateFileDialog(1, "*.eml", None, 0,
                                   "EML Files (*.eml)|*.eml|All Files (*.*)|*.*|")
    dlg.DoModal()
    return dlg.GetPathName()


def FileStartCalls():
    print(colored("\n    📄 Select an .eml file (press Cancel to abort):", "cyan"))

    try:
        filepath = SelectEMLFile()
        if not filepath or not os.path.isfile(filepath):
            print(colored("    ❌ No file selected. Returning to menu.\n", "red"))
            return

        log(f"\n[EXCEL_FILE] EML selected via explorer: {filepath}")
        print(colored(f"    ✅ Selected file: {filepath}", "green"))
        EmailSpecsFromFile(filepath)
        print(colored("    📦 All spec modules loaded successfully.\n", "cyan"))

    except Exception as e:
        print(colored(f"    ❌ Error during file selection: {e}", "red"))
        return



def FileExcelLogging():
    folder = MakeWalkerExcelFolder()
    FileStartCalls()

    template_path = GetTemplatePath()
    wb = load_workbook(template_path)
    sheet = wb.active

    field_map = {
        "E-mail sender": GLOBALS.SENDER_ADDRESS,
        "E-mail subject": GLOBALS.SENDER_SUBJECT,
        "E-mail Source IP": GLOBALS.SENDER_IP,
        "E-mail display name": GLOBALS.SENDER_DISPLAY,
        "Artifacts dropped (file)": GLOBALS.FILES,
        "Malicious URL": GLOBALS.MAL_URL,
        "Credential POST": "",
        "Redirect URL": "",
        "Research links": ""
    }

    filled_labels = set()

    for row in sheet.iter_rows(min_row=2):
        label = row[4].value
        if not label or label in filled_labels or label not in field_map:
            continue
        if row[3].value:
            continue

        value = field_map[label]
        if isinstance(value, list):
            value = "\n".join(value)
        row[3].value = value
        filled_labels.add(label)

    timestamp = datetime.now().strftime("%Y-%m-%d")
    name = input("\n    📄 Name this Excel sheet (Datetime-[INPUT].xlsx): ").strip()
    filename = f"{timestamp}-{name}.xlsx"
    save_path = os.path.join(folder, filename)
    wb.save(save_path)

    ResetTemplateSheet(sheet)
    wb.save(template_path)

    log(f"\n[EXCEL_LOGGING] Auto-filled sheet saved to: {save_path}")
    print(colored(f"\n    ✅ Excel sheet saved to: {save_path}\n", "green"))


    # TODO : CODE TO WORK ON THE OPTIONALS POST REDIRECTS AND RLs
    # TODO : ALSO NEED TO MAKE A CORRECTION / MAKE SURE THIS IS RIGHT Function to CALL TO
    # TODO : REPLACE WRONGFUL GLOBALS
    # TODO: FIX SENDER ADDRESS




def GuidedExcelLogging():
    folder = MakeWalkerExcelFolder()


    return

def ManualExcelLogging():

    folder = MakeWalkerExcelFolder()
    template_path = GetTemplatePath()
    wb = load_workbook(template_path)
    sheet = wb.active

    field_order = [
        "E-mail sender",
        "E-mail subject",
        "E-mail Source IP",
        "E-mail display name",
        "Artifacts dropped (file)",
        "Malicious URL",
        "Credential POST",
        "Redirect URL",
        "Research links"
    ]

    for label in field_order:
        print(colored(f"\n    🔹 Enter values for {label.upper()} (press Enter to skip, 'n' to cancel)", "yellow"))

        row_count = 0
        for row in sheet.iter_rows(min_row=2):
            comment = row[4].value
            if comment != label:
                continue
            if row[3].value:
                continue

            val = input("     ➤ Value: ").strip()
            if val.lower() == "n":
                ResetTemplateSheet(sheet)
                wb.save(template_path)
                print(colored("\n    ❌ Logging manually canceled by user.\n", "red"))
                return

            elif val == "":
                break
            else:
                row[3].value = val
                row_count += 1

            if row_count >= 5:
                break

    timestamp = datetime.now().strftime("%Y-%m-%d")
    name = input("\n    📄 Name this Excel sheet (Datetime-[INPUT].xlsx): ").strip()
    filename = f"{timestamp}-{name}.xlsx"
    save_path = os.path.join(folder, filename)
    wb.save(save_path)
    ResetTemplateSheet(sheet)
    wb.save(template_path)

    log(f"\n[EXCEL_LOGGING] Manual sheet saved to: {save_path}")
    print(colored(f"\n    ✅ Excel sheet saved to: {save_path}\n", "green"))

def ResetTemplateSheet(sheet):
    for row in sheet.iter_rows(min_row=2):
        if row[3].value:
            row[3].value = None


def ExcelHandlerMain():
    print(colored("\n    *-------------------------📶 Excel Logging 📶--------------------------*", "magenta"))

    mode = input("\n    Choose logging mode. Guided, Manual, or File (g/m/f): ").strip().lower()

    if mode == "g":
        GuidedExcelLogging()
    elif mode == "m":
        ManualExcelLogging()
    elif mode == "f":
        FileExcelLogging()
    else:
        print(colored("    ❌ Invalid option. Please choose 'G','M', or 'f'.", "red"))

    print(colored("\n    *----------------------------------------------------------------------*", "magenta"))